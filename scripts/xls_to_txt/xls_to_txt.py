# -*- coding: utf-8 -*
from openpyxl import load_workbook

# 表格名称、sheet名称、起始行、终止行、第几列
def get_list_data_from_xlsx(filename,sheet_name,row_start,row_end,column_num):
    # 读取路径
    book = load_workbook(filename,'r')
    # 读取名字为Sheet1的表
    sheet = book.get_sheet_by_name(sheet_name)
    row_num = row_start
    # 将表中第一列的1-100行数据写入data数组中
    while row_num <= row_end:
        raw_data = sheet.cell(row=row_num, column=column_num).value
        raw_data=round(raw_data,3)
        list_data.append(raw_data)
        row_num = row_num + 1

def write_data_to_file(filename, list_data,smoke_data_name):
    f_output=open(filename,'w')
    f_output.write('const float '+ smoke_data_name+'['+str(len(list_data))+'] = \n')
    f_output.write('{\n     ')
    string =  str(list_data)
    string=string[1:len(string)-1]
    count = 0   
    for index,value in enumerate(string):
        f_output.write(value)
        if ',' == value:
            count += 1
            if count%10==0:
                f_output.write('\n    ')
            # print(',','index:',index)#出现的位置
    # print(',','count:',count)#出现的次数
    f_output.write('\n};')
    f_output.close()

def read_data_to_file(filename):
    f_input=open(filename,'r')
    data = f_input.read()
    print(data)

xlsx_name = 'HNS103F3435F40B1392阻温对应表.xlsx' # 读取的表格名称
sheet_name = 'NTC' # 表格中的sheet名称
row_start = 7       # 起始行
row_end = 172        # 终止行
column_num=3        # 列数
list_data = []      # 用于存储数据的数组
output_f = 'ntc_temp.h'
array_name='res_array'
get_list_data_from_xlsx(xlsx_name,sheet_name,row_start,row_end,column_num)
write_data_to_file(output_f, list_data,array_name)
# print(list_data)
read_data_to_file(output_f)

